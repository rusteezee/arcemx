"""Import INDmoney "Consolidated Tax Report" XLSX files into Supabase.

INDmoney's tax report (LTCG + STCG sheets) is the only source of truth
for realized capital gains/losses - the account's own transactions
ledger (see import_indmoney_transactions.py) tracks buys/sells for the
Value Timeline replay, but doesn't classify gains by holding period or
compute the tax-relevant figures (expense, disallowed losses u/s 94(8),
taxable gain). This importer reads those sheets directly and loads one
row per sold lot into `realized_pnl`.

Usage:
    python -m fetchers.import_realized_pnl <file1.xlsx> [file2.xlsx ...]
    python -m fetchers.import_realized_pnl W:/consolidated_tax_report_*.xlsx

Deliberately does NOT use pandas.read_excel with a fixed header row -
each sheet has 7 different asset-category sub-tables (Indian Stocks,
Equity ETFs, Equity Mutual Funds, Non-Equity ETFs, Non-Equity Mutual
Funds, US Stocks, Rights Entitlements), each with its own column layout,
a two-row header (a merged label row + a sub-header row for split
Redemption/Purchase columns), and a Gains/Losses/Total row structure.
Row OFFSETS shift between report years (INDmoney added a "Buybacks" row
to the summary table between FY2025-26 and FY2026-27's report), so every
section is located by searching for its label text, never a fixed row
number - the layout logic will survive next year's report format
drifting again, as long as the labels themselves don't change.

Re-running the importer is safe: rows are upserted against a composite
key (user_id, fy, gain_type, asset_category, isin, sell_date, buy_date,
units_sold, sell_value), so repeated runs on the same file won't
duplicate rows. A financial year with a placeholder "-" in every data
row (nothing sold that year, or in that asset category) imports 0 rows
for that section, which is correct, not a failure.
"""
from __future__ import annotations

import os
import sys
import glob
import re
from datetime import datetime
from typing import Iterable

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv()

# Section label -> asset_category value stored in the DB. Matched via a
# normalised (lowercased, whitespace/dash-collapsed) comparison so label
# spacing drift between report years ("Non-Equity ETFs" vs "Non - Equity
# ETFs", both observed across FY2025-26 vs FY2026-27) doesn't break the
# match.
SECTION_LABELS = {
    "indian stocks": "Indian Stocks",
    "equity etfs": "Equity ETFs",
    "equity mutual funds": "Equity Mutual Funds",
    "non equity etfs": "Non-Equity ETFs",
    "non equity mutual funds": "Non-Equity Mutual Funds",
    "us stocks": "US Stocks",
    "rights entitlements": "Rights Entitlements",
}


def _norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[\s\-]+", " ", s)
    return s


def _sb_client():
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    if not url or not key:
        raise SystemExit(
            "SUPABASE_URL / SUPABASE_KEY env vars are required. "
            "Set them in your shell before running this importer."
        )
    return create_client(url, key)


def _num(v):
    """Best-effort numeric coercion. INDmoney fills empty cells with a
    literal '-' string, not a blank cell - treat that as 0/None."""
    if v is None or v == "-":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _date(v):
    if v is None or v == "-":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    if not s or s == "-":
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _find_section_starts(ws) -> list[tuple[int, str]]:
    """Scan column A/B (report layout shifts LTCG/STCG's start column by
    one) for section label cells. Returns [(row_index, asset_category)]
    in document order."""
    hits: list[tuple[int, str]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row[:3]:
            key = _norm(cell.value)
            if key in SECTION_LABELS:
                hits.append((cell.row, SECTION_LABELS[key]))
                break
    return hits


def _find_header_row(ws, start_row: int, end_row: int) -> int | None:
    """Within [start_row, end_row), find the row containing 'Name of
    Stock' / 'Name of Mutual Fund' - the section's actual data header."""
    for r in range(start_row, end_row):
        for cell in ws[r]:
            v = _norm(cell.value)
            if v in ("name of stock", "name of mutual fund"):
                return r
    return None


def _row_values(ws, r: int) -> list:
    return [c.value for c in ws[r]]


def _is_placeholder_row(vals: list) -> bool:
    """A row of all '-' (or all-empty) means 'nothing here', not data."""
    nonblank = [v for v in vals if v is not None]
    if not nonblank:
        return True
    return all(v == "-" for v in nonblank)


def _extract_standard_section(ws, header_row: int, next_boundary: int,
                                gain_type: str, asset_category: str,
                                fy: str, source_file: str) -> list[dict]:
    """Indian Stocks / Equity ETFs / Non-Equity ETFs / Rights
    Entitlements share the same core layout: Name | ISIN | dates |
    Holding Period | Units | Per-unit x2 | Total x2 | [FMV+CoA, LTCG
    only] | Gross Gain | Expense | Disallowed Losses | Taxable Gain.
    Located by column HEADER TEXT, not position, so the LTCG-only extra
    columns (FMV as on 31 Jan 2018, Cost of Acquisition) don't shift
    anything for STCG sections.
    """
    header_cells = ws[header_row]
    col_of: dict[str, int] = {}
    for cell in header_cells:
        label = _norm(cell.value)
        if label and label not in col_of:
            col_of[label] = cell.column
    # Sub-header row (redemption/purchase split for 'per unit' + 'total')
    sub_row = header_row + 1
    sub_cells = ws[sub_row]
    for cell in sub_cells:
        label = _norm(cell.value)
        if label in ("redemption value", "purchase value") and cell.column not in col_of.values():
            # Disambiguate: first occurrence pair = per-unit, second = total.
            pass  # handled positionally below; header row alone is ambiguous for split cols

    def col(label: str) -> int | None:
        return col_of.get(label)

    name_c = col("name of stock")
    isin_c = col("isin")
    redemption_date_c = col("redemption date")
    purchase_date_c = col("purchase date")
    holding_c = col("holding period")
    units_c = col("units sold")
    gross_c = col("gross gains/losses")
    expense_c = col("expense")
    taxable_c = col("taxable gains/losses")

    if name_c is None or units_c is None or taxable_c is None:
        return []

    # "Per unit" spans 2 columns (redemption, purchase) starting right
    # after Units sold; "Total" spans the next 2. Positional from
    # units_c since the split sub-header can't be disambiguated by text
    # alone (both pairs say "Redemption Value" / "Purchase Value").
    per_unit_redemption_c = units_c + 1
    per_unit_purchase_c = units_c + 2
    total_redemption_c = units_c + 3
    total_purchase_c = units_c + 4

    out = []
    in_data_block = False
    for r in range(header_row + 2, next_boundary):
        first_cell = ws.cell(row=r, column=1).value or ws.cell(row=r, column=2).value
        label = _norm(first_cell)
        if label in ("gains", "losses"):
            in_data_block = True
            continue
        if label == "total":
            in_data_block = False
            continue
        if not in_data_block:
            continue
        vals = _row_values(ws, r)
        if _is_placeholder_row(vals[:units_c + 5] if units_c else vals):
            continue
        name = ws.cell(row=r, column=name_c).value
        if not name or name == "-":
            continue
        out.append({
            "fy": fy,
            "gain_type": gain_type,
            "asset_category": asset_category,
            "scrip_name": str(name).strip(),
            "isin": _nan_none(ws.cell(row=r, column=isin_c).value) if isin_c else None,
            "sell_date": _date(ws.cell(row=r, column=redemption_date_c).value) if redemption_date_c else None,
            "buy_date": _date(ws.cell(row=r, column=purchase_date_c).value) if purchase_date_c else None,
            "holding_period": _nan_none(ws.cell(row=r, column=holding_c).value) if holding_c else None,
            "units_sold": _num(ws.cell(row=r, column=units_c).value),
            "sell_value": _num(ws.cell(row=r, column=total_redemption_c).value),
            "buy_value": _num(ws.cell(row=r, column=total_purchase_c).value),
            "gross_gain_loss": _num(ws.cell(row=r, column=gross_c).value) if gross_c else None,
            "expense": _num(ws.cell(row=r, column=expense_c).value) if expense_c else None,
            "taxable_gain_loss": _num(ws.cell(row=r, column=taxable_c).value),
            "currency": "INR",
            "source_file": source_file,
        })
    return out


def _nan_none(v):
    if v is None or v == "-":
        return None
    return str(v).strip()


def _read_sheet(ws, gain_type: str, fy: str, source_file: str) -> list[dict]:
    sections = _find_section_starts(ws)
    if not sections:
        return []
    boundaries = [s[0] for s in sections[1:]] + [ws.max_row + 1]
    rows: list[dict] = []
    for (start_row, asset_category), end_row in zip(sections, boundaries):
        if asset_category in ("Equity Mutual Funds", "Non-Equity Mutual Funds", "US Stocks"):
            # Distinct column layouts (Folio No. / USD+INR dual currency).
            # Not built yet - no real data to validate against. Skipped
            # cleanly rather than guessing the mapping wrong.
            continue
        header_row = _find_header_row(ws, start_row, end_row)
        if header_row is None:
            continue
        rows.extend(_extract_standard_section(
            ws, header_row, end_row, gain_type, asset_category, fy, source_file
        ))
    return rows


def _fy_from_filename(path: str) -> str:
    m = re.search(r"(\d{4})[-_](\d{2,4})", os.path.basename(path))
    if m:
        return f"{m.group(1)}-{m.group(2)[-2:]}"
    return "unknown"


def import_files(paths: list[str], user_id: str = "default") -> dict:
    sb = _sb_client()
    total_loaded = 0
    per_file: dict[str, int] = {}

    for path in paths:
        if not os.path.exists(path):
            print(f"[skip] {path}: file not found")
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            print(f"[skip] {path}: read failed. {e!r}")
            continue

        fy = _fy_from_filename(path)
        source_file = os.path.basename(path)
        records: list[dict] = []
        for sheet_name, gain_type in (("LTCG", "LTCG"), ("STCG", "STCG")):
            if sheet_name not in wb.sheetnames:
                continue
            recs = _read_sheet(wb[sheet_name], gain_type, fy, source_file)
            for rec in recs:
                rec["user_id"] = user_id
            records.extend(recs)

        if not records:
            print(f"[empty] {source_file}: 0 realized rows (FY {fy})")
            per_file[path] = 0
            continue

        chunk = 100
        loaded = 0
        for i in range(0, len(records), chunk):
            batch = records[i:i + chunk]
            try:
                sb.table("realized_pnl").upsert(
                    batch,
                    on_conflict="user_id,fy,gain_type,asset_category,isin,sell_date,buy_date,units_sold,sell_value",
                ).execute()
                loaded += len(batch)
            except Exception as e:
                print(f"  [batch fail] {source_file} rows {i}-{i+len(batch)}: {e!r}")
        per_file[path] = loaded
        total_loaded += loaded
        print(f"[ok]   {source_file}: {loaded} rows upserted (FY {fy})")

    print("\n" + "=" * 60)
    print(f"Imported: {total_loaded} rows across {len(per_file)} file(s)")
    return {"loaded": total_loaded, "per_file": per_file}


def main():
    args = sys.argv[1:]
    if not args:
        print("Usage: python -m fetchers.import_realized_pnl <file.xlsx> [file2.xlsx ...]")
        sys.exit(2)
    user_id = os.getenv("TELEGRAM_CHAT_ID", "default")
    paths: list[str] = []
    for a in args:
        matched = glob.glob(a)
        paths.extend(matched if matched else [a])
    paths = sorted(set(paths))
    if not paths:
        print("No matching files.")
        sys.exit(2)
    print(f"Importing {len(paths)} file(s) for user_id={user_id}:")
    for p in paths:
        print(f"  - {p}")
    print()
    import_files(paths, user_id=user_id)


if __name__ == "__main__":
    main()
